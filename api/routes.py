from __future__ import annotations

import io
from typing import Any

import pandas as pd
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from engine.backtest import BacktestResult, run_backtest
from engine.data_loader import load_ohlc
from engine.params import StrategyParams

router = APIRouter(prefix="/api")


def _parse_bool(value: str | bool) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _serialize_timestamp(value: Any) -> str:
    ts = pd.Timestamp(value)
    if pd.isna(ts):
        return ""
    return ts.isoformat()


def serialize_result(result: BacktestResult) -> dict[str, Any]:
    trades = [
        {
            "trade_number": t.trade_number,
            "direction": t.direction,
            "entry_time": _serialize_timestamp(t.entry_time),
            "exit_time": _serialize_timestamp(t.exit_time),
            "entry_price": float(t.entry_price),
            "exit_price": float(t.exit_price),
            "entry_atr": float(t.entry_atr),
            "atr_result": float(t.atr_result),
            "bars_held": int(t.bars_held),
            "exit_reason": t.exit_reason,
        }
        for t in result.trades
    ]
    monthly = [
        {"month": row["month"], "net_atr": float(row["net_atr"])}
        for row in result.monthly_performance
    ]
    statistics = {}
    for key, value in result.statistics.items():
        if key == "number_of_trades":
            statistics[key] = int(value)
        elif isinstance(value, (int, float)):
            statistics[key] = float(value)
        else:
            statistics[key] = value

    return {
        "statistics": statistics,
        "trades": trades,
        "equity_curve": [float(v) for v in result.equity_curve],
        "monthly_performance": monthly,
    }


def _build_params(
    entry_lookback: int,
    exit_lookback: int,
    atr_length: int,
    use_filter: str,
    use_ema_filter: str,
    ema_length: int,
    buffer: float,
) -> StrategyParams:
    if entry_lookback < 1 or exit_lookback < 1:
        raise HTTPException(status_code=400, detail="Lookback values must be >= 1.")
    if atr_length < 2 or ema_length < 2:
        raise HTTPException(status_code=400, detail="ATR/EMA length must be >= 2.")
    if buffer < 0:
        raise HTTPException(status_code=400, detail="Buffer must be >= 0.")

    return StrategyParams(
        entry_lookback=int(entry_lookback),
        exit_lookback=int(exit_lookback),
        atr_length=int(atr_length),
        use_filter=_parse_bool(use_filter),
        use_ema_filter=_parse_bool(use_ema_filter),
        ema_length=int(ema_length),
        buffer=float(buffer),
    )


async def _run_from_upload(
    file: UploadFile,
    entry_lookback: int,
    exit_lookback: int,
    atr_length: int,
    use_filter: str,
    use_ema_filter: str,
    ema_length: int,
    buffer: float,
) -> BacktestResult:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Please upload a CSV file first.")
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are supported.")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    params = _build_params(
        entry_lookback=entry_lookback,
        exit_lookback=exit_lookback,
        atr_length=atr_length,
        use_filter=use_filter,
        use_ema_filter=use_ema_filter,
        ema_length=ema_length,
        buffer=buffer,
    )

    try:
        df = load_ohlc(io.BytesIO(raw))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # pragma: no cover - defensive parse guard
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {exc}") from exc

    if df.empty:
        raise HTTPException(status_code=400, detail="CSV produced no usable OHLC rows.")

    return run_backtest(df, params)


def _excel_from_payload(payload: dict[str, Any]) -> bytes:
    stats = payload.get("statistics", {})
    trades = payload.get("trades", [])
    monthly = payload.get("monthly_performance", [])
    equity = payload.get("equity_curve", [])

    wb = Workbook()

    ws_stats = wb.active
    ws_stats.title = "Statistics"
    ws_stats.append(["Metric", "Value"])
    for key, value in stats.items():
        ws_stats.append([key, value])

    ws_trades = wb.create_sheet("Trade log")
    ws_trades.append([
        "Trade Number",
        "Direction",
        "Entry Date",
        "Exit Date",
        "Entry Price",
        "Exit Price",
        "Entry ATR",
        "Result (ATR)",
        "Bars Held",
        "Exit Reason",
    ])
    for t in trades:
        ws_trades.append([
            t.get("trade_number"),
            t.get("direction"),
            t.get("entry_time"),
            t.get("exit_time"),
            t.get("entry_price"),
            t.get("exit_price"),
            t.get("entry_atr"),
            t.get("atr_result"),
            t.get("bars_held"),
            t.get("exit_reason"),
        ])

    ws_monthly = wb.create_sheet("Monthly performance")
    ws_monthly.append(["Month", "Net ATR"])
    for row in monthly:
        ws_monthly.append([row.get("month"), row.get("net_atr")])

    ws_equity = wb.create_sheet("Equity curve")
    ws_equity.append(["Trade", "Equity (ATR)"])
    for idx, value in enumerate(equity, start=1):
        ws_equity.append([idx, value])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@router.post("/backtest")
async def backtest(
    file: UploadFile = File(...),
    entry_lookback: int = Form(20),
    exit_lookback: int = Form(10),
    atr_length: int = Form(14),
    use_filter: str = Form("false"),
    use_ema_filter: str = Form("false"),
    ema_length: int = Form(200),
    buffer: float = Form(0.0),
) -> dict[str, Any]:
    result = await _run_from_upload(
        file=file,
        entry_lookback=entry_lookback,
        exit_lookback=exit_lookback,
        atr_length=atr_length,
        use_filter=use_filter,
        use_ema_filter=use_ema_filter,
        ema_length=ema_length,
        buffer=buffer,
    )
    return serialize_result(result)


@router.post("/export")
async def export(
    file: UploadFile = File(...),
    entry_lookback: int = Form(20),
    exit_lookback: int = Form(10),
    atr_length: int = Form(14),
    use_filter: str = Form("false"),
    use_ema_filter: str = Form("false"),
    ema_length: int = Form(200),
    buffer: float = Form(0.0),
) -> StreamingResponse:
    result = await _run_from_upload(
        file=file,
        entry_lookback=entry_lookback,
        exit_lookback=exit_lookback,
        atr_length=atr_length,
        use_filter=use_filter,
        use_ema_filter=use_ema_filter,
        ema_length=ema_length,
        buffer=buffer,
    )
    payload = serialize_result(result)
    content = _excel_from_payload(payload)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="turtle_backtest_results.xlsx"'
        },
    )
